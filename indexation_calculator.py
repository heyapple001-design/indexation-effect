# indexation_calculator.py (финальная версия с правильным кешированием)
import sys
import os
import pandas as pd
import numpy as np
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import warnings
from datetime import datetime
warnings.filterwarnings('ignore')

# Импортируем маппинги
from mappings import *

class ColumnSelectionDialog(QDialog):
    """Диалог для выбора колонок из файла"""
    def __init__(self, columns, title="Выбор колонок", parent=None, required_only=False):
        super().__init__(parent)
        self.columns = columns
        self.title = title
        self.required_only = required_only
        self.selected = {}
        self.initUI()
        
    def initUI(self):
        self.setWindowTitle(self.title)
        self.setModal(True)
        layout = QVBoxLayout()
        
        scroll = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        
        self.comboboxes = {}
        
        if self.required_only:
            # Для файла начислений - только MSISDN и сумма
            labels = [
                ("MSISDN (идентификатор)", "msisdn"),
                ("Сумма начислений", "charges")
            ]
        else:
            # Для базы индексации - все колонки
            labels = [
                ("MSISDN (идентификатор)", "msisdn"),
                ("Начисления ДО индексации", "charges_before"),
                ("Краткий сегмент", "short_seg"),
                ("Подсегмент", "sub_seg"),
                ("МРФ", "mrf"),
                ("РФ (регион)", "region"),
                ("Процент индексации", "percent")
            ]
        
        for label, key in labels:
            row_layout = QHBoxLayout()
            row_layout.addWidget(QLabel(label + ":"))
            combo = QComboBox()
            combo.addItem("-- Не выбрано --")
            combo.addItems(self.columns)
            combo.setCurrentIndex(0)
            self.comboboxes[key] = combo
            row_layout.addWidget(combo)
            scroll_layout.addLayout(row_layout)
            
        scroll.setWidget(scroll_widget)
        layout.addWidget(scroll)
        
        # Кнопки
        btn_layout = QHBoxLayout()
        ok_btn = QPushButton("OK")
        ok_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("Отмена")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(ok_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)
        
        self.setLayout(layout)
        self.resize(500, 400)
        
    def get_selections(self):
        """Возвращает выбранные колонки"""
        result = {}
        for key, combo in self.comboboxes.items():
            value = combo.currentText()
            if value != "-- Не выбрано --":
                result[key] = value
        return result

class DateSelectionDialog(QDialog):
    """Диалог для выбора даты индексации"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.selected_date = None
        self.initUI()
        
    def initUI(self):
        self.setWindowTitle("Выбор даты индексации")
        self.setModal(True)
        layout = QVBoxLayout()
        
        self.calendar = QCalendarWidget()
        self.calendar.setGridVisible(True)
        layout.addWidget(self.calendar)
        
        btn_layout = QHBoxLayout()
        ok_btn = QPushButton("OK")
        ok_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("Отмена")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(ok_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)
        
        self.setLayout(layout)
        self.resize(400, 350)
        
    def accept(self):
        self.selected_date = self.calendar.selectedDate().toPyDate()
        super().accept()
        
    def get_date(self):
        return self.selected_date

class IndexationApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.files_loaded = {
            'report': None,
            'base': []
        }
        # Кеш для текущего МРФ
        self.cache = {
            'charges_df': None,          # DataFrame начислений для текущего МРФ
            'charges_dict': None,        # Словарь MSISDN->сумма для текущего МРФ
            'charges_file': None,        # Имя файла
            'charges_columns': None,     # Выбранные колонки
            'charges_loaded': False,     # Флаг загрузки
            'charges_count': 0,          # Количество записей
            'current_mrf': None          # Для какого МРФ загружены начисления
        }
        self.data = {
            'report_df': None,
            'base_df': None,
            'selected_mrf': [],
            'selected_regions': [],
            'selected_month': None,
            'selected_service': None,
            'indexation_date': None,
            'indexation_percent': None,
            'selected_segments': [],
            'selected_subsegments': [],
            'mode': None,
            'column_mapping': {},
            'service_name': None,
            'current_mrf_index': 0,
            'processing_mrf': None       # Какой МРФ сейчас обрабатывается
        }
        self.results = {}
        self.log = []
        
    def initUI(self):
        self.setWindowTitle('Расчет эффекта индексации')
        self.setGeometry(100, 100, 1300, 900)
        
        # Центральный виджет
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        
        # Создаем вкладки
        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs)
        
        # Вкладка 1: Настройка
        self.setup_tab = QWidget()
        self.tabs.addTab(self.setup_tab, "Настройка")
        self.setup_ui()
        
        # Вкладка 2: Лог
        self.log_tab = QWidget()
        self.tabs.addTab(self.log_tab, "Лог")
        self.log_ui()
        
        # Вкладка 3: Результаты
        self.results_tab = QWidget()
        self.tabs.addTab(self.results_tab, "Результаты")
        self.results_ui()
        
        self.add_log("Программа запущена")
        
    def setup_ui(self):
        layout = QVBoxLayout(self.setup_tab)
        
        # Создаем скролл-зону
        scroll = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        
        # 1. Выбор месяца
        month_group = QGroupBox("1. Выбор отчетного месяца")
        month_layout = QHBoxLayout()
        self.month_combo = QComboBox()
        self.month_combo.addItems(MONTHS)
        self.month_combo.currentTextChanged.connect(self.on_month_changed)
        month_layout.addWidget(QLabel("Отчетный месяц:"))
        month_layout.addWidget(self.month_combo)
        month_layout.addStretch()
        month_group.setLayout(month_layout)
        scroll_layout.addWidget(month_group)
        
        # 2. Выбор МРФ
        mrf_group = QGroupBox("2. Выбор макрорегионов")
        mrf_layout = QVBoxLayout()
        
        mrf_grid = QGridLayout()
        self.mrf_checkboxes = {}
        row, col = 0, 0
        for mrf in ALL_MRF:
            checkbox = QCheckBox(mrf)
            checkbox.stateChanged.connect(self.on_mrf_changed)
            self.mrf_checkboxes[mrf] = checkbox
            mrf_grid.addWidget(checkbox, row, col)
            col += 1
            if col > 3:
                col = 0
                row += 1
        mrf_layout.addLayout(mrf_grid)
        
        select_all_btn = QPushButton("Выбрать все")
        select_all_btn.clicked.connect(self.select_all_mrf)
        mrf_layout.addWidget(select_all_btn)
        
        mrf_group.setLayout(mrf_layout)
        scroll_layout.addWidget(mrf_group)
        
        # 3. Выбранные регионы
        regions_group = QGroupBox("Выбранные регионы")
        regions_layout = QVBoxLayout()
        self.regions_text = QTextEdit()
        self.regions_text.setReadOnly(True)
        self.regions_text.setMaximumHeight(100)
        regions_layout.addWidget(self.regions_text)
        regions_group.setLayout(regions_layout)
        scroll_layout.addWidget(regions_group)
        
        # 4. Загрузка файлов
        files_group = QGroupBox("3. Загрузка файлов")
        files_layout = QVBoxLayout()
        
        btn_layout = QHBoxLayout()
        self.load_report_btn = QPushButton("Загрузить файл отчета")
        self.load_report_btn.clicked.connect(self.load_report_file)
        self.load_report_btn.setStyleSheet("background-color: #e1f5fe; padding: 5px;")
        
        self.load_base_btn = QPushButton("Загрузить базу индексации")
        self.load_base_btn.clicked.connect(self.load_base_file)
        self.load_base_btn.setStyleSheet("background-color: #fff3e0; padding: 5px;")
        
        self.load_charges_btn = QPushButton("Загрузить начисления (режим А)")
        self.load_charges_btn.clicked.connect(self.load_charges_file)
        self.load_charges_btn.setStyleSheet("background-color: #e8f5e9; padding: 5px;")
        self.load_charges_btn.setEnabled(False)
        
        btn_layout.addWidget(self.load_report_btn)
        btn_layout.addWidget(self.load_base_btn)
        btn_layout.addWidget(self.load_charges_btn)
        files_layout.addLayout(btn_layout)
        
        # Информация о файлах
        self.files_info = QTextEdit()
        self.files_info.setReadOnly(True)
        self.files_info.setMaximumHeight(80)
        files_layout.addWidget(self.files_info)
        
        # Информация о кеше начислений для текущего МРФ
        cache_group = QGroupBox("Кеш начислений (текущий МРФ)")
        cache_layout = QVBoxLayout()
        
        cache_status_layout = QHBoxLayout()
        self.cache_status_label = QLabel("❌ Не загружены")
        self.cache_status_label.setStyleSheet("color: red; font-weight: bold;")
        self.cache_info_label = QLabel("")
        self.clear_cache_btn = QPushButton("Очистить кеш")
        self.clear_cache_btn.clicked.connect(self.clear_cache)
        self.clear_cache_btn.setEnabled(False)
        
        cache_status_layout.addWidget(QLabel("Статус:"))
        cache_status_layout.addWidget(self.cache_status_label)
        cache_status_layout.addWidget(self.cache_info_label)
        cache_status_layout.addStretch()
        cache_status_layout.addWidget(self.clear_cache_btn)
        cache_layout.addLayout(cache_status_layout)
        
        # Информация о текущем МРФ
        self.cache_mrf_label = QLabel("Текущий МРФ: не выбран")
        cache_layout.addWidget(self.cache_mrf_label)
        
        cache_group.setLayout(cache_layout)
        files_layout.addWidget(cache_group)
        
        files_group.setLayout(files_layout)
        scroll_layout.addWidget(files_group)
        
        # 5. Выбор режима
        mode_group = QGroupBox("4. Выбор режима")
        mode_layout = QHBoxLayout()
        self.mode_a_radio = QRadioButton("Режим А: Начисления в отдельном файле")
        self.mode_b_radio = QRadioButton("Режим Б: Начисления в базе индексации")
        self.mode_a_radio.toggled.connect(self.on_mode_changed)
        self.mode_a_radio.setChecked(True)
        
        mode_layout.addWidget(self.mode_a_radio)
        mode_layout.addWidget(self.mode_b_radio)
        mode_layout.addStretch()
        mode_group.setLayout(mode_layout)
        scroll_layout.addWidget(mode_group)
        
        # 6. Настройка колонок
        columns_group = QGroupBox("5. Настройка колонок")
        columns_layout = QVBoxLayout()
        
        self.column_setup_btn = QPushButton("Настроить соответствие колонок в базе индексации")
        self.column_setup_btn.clicked.connect(self.setup_columns)
        self.column_setup_btn.setEnabled(False)
        columns_layout.addWidget(self.column_setup_btn)
        
        self.column_info = QLabel("Колонки не настроены")
        self.column_info.setStyleSheet("color: gray;")
        columns_layout.addWidget(self.column_info)
        
        columns_group.setLayout(columns_layout)
        scroll_layout.addWidget(columns_group)
        
        # 7. Дополнительные параметры
        params_group = QGroupBox("6. Дополнительные параметры")
        params_layout = QGridLayout()
        
        # Дата индексации
        params_layout.addWidget(QLabel("Дата индексации:"), 0, 0)
        self.date_btn = QPushButton("Выбрать дату")
        self.date_btn.clicked.connect(self.select_date)
        self.date_btn.setEnabled(False)
        params_layout.addWidget(self.date_btn, 0, 1)
        self.date_label = QLabel("Не выбрана")
        params_layout.addWidget(self.date_label, 0, 2)
        
        # Услуга
        params_layout.addWidget(QLabel("Услуга:"), 1, 0)
        self.service_combo = QComboBox()
        self.service_combo.addItem("-- Выберите услугу --")
        self.service_combo.addItems(SERVICES)
        self.service_combo.setEnabled(False)
        params_layout.addWidget(self.service_combo, 1, 1, 1, 2)
        
        # Процент индексации
        params_layout.addWidget(QLabel("Процент индексации (%):"), 2, 0)
        self.percent_edit = QLineEdit()
        self.percent_edit.setPlaceholderText("Например: 10")
        self.percent_edit.setEnabled(False)
        params_layout.addWidget(self.percent_edit, 2, 1, 1, 2)
        
        # Выбор сегментов
        params_layout.addWidget(QLabel("Краткие сегменты:"), 3, 0)
        self.short_seg_list = QListWidget()
        self.short_seg_list.setSelectionMode(QListWidget.MultiSelection)
        for seg in SHORT_SEGMENTS:
            self.short_seg_list.addItem(seg)
        self.short_seg_list.setEnabled(False)
        params_layout.addWidget(self.short_seg_list, 3, 1, 1, 2)
        
        params_layout.addWidget(QLabel("Подсегменты:"), 4, 0)
        self.sub_seg_list = QListWidget()
        self.sub_seg_list.setSelectionMode(QListWidget.MultiSelection)
        for seg in SUB_SEGMENTS:
            self.sub_seg_list.addItem(seg)
        self.sub_seg_list.setEnabled(False)
        params_layout.addWidget(self.sub_seg_list, 4, 1, 1, 2)
        
        params_group.setLayout(params_layout)
        scroll_layout.addWidget(params_group)
        
        # Кнопка запуска
        self.run_btn = QPushButton("▶ Запустить расчет")
        self.run_btn.clicked.connect(self.run_calculation)
        self.run_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-size: 14px;
                padding: 10px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
        """)
        self.run_btn.setEnabled(False)
        scroll_layout.addWidget(self.run_btn)
        
        # Прогресс-бар
        self.progress_bar = QProgressBar()
        scroll_layout.addWidget(self.progress_bar)
        
        scroll_layout.addStretch()
        scroll.setWidget(scroll_widget)
        scroll.setWidgetResizable(True)
        layout.addWidget(scroll)
        
    def log_ui(self):
        layout = QVBoxLayout(self.log_tab)
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        layout.addWidget(self.log_text)
        
    def results_ui(self):
        layout = QVBoxLayout(self.results_tab)
        self.results_text = QTextEdit()
        self.results_text.setReadOnly(True)
        layout.addWidget(self.results_text)
        
    def add_log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}"
        self.log.append(log_entry)
        if hasattr(self, 'log_text'):
            self.log_text.append(log_entry)
        print(log_entry)
        
    def update_cache_status(self):
        """Обновляет статус кеша начислений для текущего МРФ"""
        if self.cache['charges_loaded']:
            self.cache_status_label.setText("✅ Загружены")
            self.cache_status_label.setStyleSheet("color: green; font-weight: bold;")
            mrf_name = self.cache['current_mrf'] if self.cache['current_mrf'] else "неизвестный"
            self.cache_info_label.setText(
                f"Файл: {os.path.basename(self.cache['charges_file'])} | "
                f"Записей: {self.cache['charges_count']:,} | "
                f"МРФ: {mrf_name}"
            )
            self.clear_cache_btn.setEnabled(True)
        else:
            self.cache_status_label.setText("❌ Не загружены")
            self.cache_status_label.setStyleSheet("color: red; font-weight: bold;")
            self.cache_info_label.setText("")
            self.clear_cache_btn.setEnabled(False)
            
    def clear_cache(self):
        """Очищает кеш начислений для текущего МРФ"""
        if self.cache['charges_loaded']:
            mrf_name = self.cache['current_mrf'] if self.cache['current_mrf'] else "неизвестный"
            self.add_log(f"Очистка кеша начислений для МРФ: {mrf_name}")
        
        self.cache['charges_df'] = None
        self.cache['charges_dict'] = None
        self.cache['charges_file'] = None
        self.cache['charges_columns'] = None
        self.cache['charges_loaded'] = False
        self.cache['charges_count'] = 0
        self.cache['current_mrf'] = None
        self.update_cache_status()
        
    def on_month_changed(self, month):
        self.data['selected_month'] = month
        self.add_log(f"Выбран месяц: {month}")
        self.check_ready()
        
    def on_mrf_changed(self):
        selected_mrf = []
        for mrf, checkbox in self.mrf_checkboxes.items():
            if checkbox.isChecked():
                selected_mrf.append(mrf)
        
        self.data['selected_mrf'] = selected_mrf
        
        regions_text = ""
        for mrf in selected_mrf:
            regions = MRF_TO_RF.get(mrf, [])
            regions_text += f"\n{mrf}:\n"
            for region in regions:
                regions_text += f"  - {region}\n"
        
        self.regions_text.setText(regions_text)
        self.add_log(f"Выбраны МРФ: {', '.join(selected_mrf)}")
        self.check_ready()
        
    def select_all_mrf(self):
        for checkbox in self.mrf_checkboxes.values():
            checkbox.setChecked(True)
            
    def load_report_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Выберите файл отчета", "", 
            "Excel files (*.xlsx);;All files (*.*)"
        )
        if file_path:
            try:
                self.data['report_df'] = pd.read_excel(file_path, sheet_name='Факт-эффект индексации 2026')
                self.files_loaded['report'] = file_path
                self.update_files_info()
                self.add_log(f"Загружен файл отчета: {os.path.basename(file_path)}")
                self.add_log(f"Колонки в отчете: {list(self.data['report_df'].columns)}")
                self.check_ready()
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить файл отчета:\n{str(e)}")
                self.add_log(f"Ошибка загрузки отчета: {str(e)}")
                
    def load_base_file(self):
        if self.files_loaded['report'] is None:
            QMessageBox.warning(self, "Предупреждение", "Сначала загрузите файл отчета!")
            return
            
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Выберите базу индексации", "", 
            "Excel files (*.xlsx);;All files (*.*)"
        )
        if file_path:
            try:
                df = pd.read_excel(file_path, sheet_name='Лист1')
                self.files_loaded['base'].append(file_path)
                if self.data['base_df'] is None:
                    self.data['base_df'] = df
                else:
                    self.data['base_df'] = pd.concat([self.data['base_df'], df], ignore_index=True)
                self.update_files_info()
                self.add_log(f"Загружена база индексации: {os.path.basename(file_path)}")
                self.add_log(f"Колонки в базе: {list(df.columns)}")
                
                # Включаем настройку колонок
                self.column_setup_btn.setEnabled(True)
                self.date_btn.setEnabled(True)
                self.service_combo.setEnabled(True)
                self.percent_edit.setEnabled(True)
                self.short_seg_list.setEnabled(True)
                self.sub_seg_list.setEnabled(True)
                
                # Спрашиваем о загрузке еще одной базы
                reply = QMessageBox.question(
                    self, "Загрузка базы", 
                    "Загрузить еще одну базу индексации?",
                    QMessageBox.Yes | QMessageBox.No
                )
                if reply == QMessageBox.Yes:
                    self.load_base_file()
                else:
                    self.check_ready()
                    
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить базу индексации:\n{str(e)}")
                self.add_log(f"Ошибка загрузки базы: {str(e)}")
                
    def load_charges_file(self):
        """Загружает файл начислений для текущего МРФ"""
        if self.files_loaded['report'] is None:
            QMessageBox.warning(self, "Предупреждение", "Сначала загрузите файл отчета!")
            return
            
        # Проверяем, для какого МРФ загружаем начисления
        if not self.data['selected_mrf']:
            QMessageBox.warning(self, "Предупреждение", "Сначала выберите МРФ!")
            return
            
        # Определяем текущий МРФ
        current_mrf = self.data.get('processing_mrf')
        if current_mrf is None and self.data['selected_mrf']:
            current_mrf = self.data['selected_mrf'][0]  # Берем первый МРФ для загрузки
            
        if not current_mrf:
            QMessageBox.warning(self, "Предупреждение", "Не удалось определить МРФ для загрузки начислений!")
            return
            
        # Проверяем, не загружены ли уже начисления для этого МРФ
        if self.cache['charges_loaded'] and self.cache['current_mrf'] == current_mrf:
            reply = QMessageBox.question(
                self, 
                "Кеш начислений", 
                f"Начисления для МРФ {current_mrf} уже загружены.\n"
                f"Файл: {os.path.basename(self.cache['charges_file'])}\n"
                f"Записей: {self.cache['charges_count']:,}\n\n"
                "Загрузить новый файл? (текущий кеш будет очищен)",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.No:
                return
            else:
                self.clear_cache()
            
        file_path, _ = QFileDialog.getOpenFileName(
            self, 
            f"Выберите файл начислений для МРФ {current_mrf}", 
            "", 
            "Excel files (*.xlsx);;All files (*.*)"
        )
        if file_path:
            try:
                # Загружаем DataFrame
                df = pd.read_excel(file_path, sheet_name='Sheet')
                
                # Пользователь выбирает колонки
                columns = list(df.columns)
                dialog = ColumnSelectionDialog(columns, 
                    f"Выбор колонок в файле начислений для МРФ {current_mrf}", 
                    self, 
                    required_only=True
                )
                if dialog.exec_() != QDialog.Accepted:
                    return
                    
                charges_mapping = dialog.get_selections()
                if 'msisdn' not in charges_mapping or 'charges' not in charges_mapping:
                    QMessageBox.warning(self, "Ошибка", "Не выбраны обязательные колонки!")
                    return
                    
                # Переименовываем колонки
                df_renamed = df.rename(columns={
                    charges_mapping['msisdn']: 'MSISDN_charges',
                    charges_mapping['charges']: 'AP_POSLE'
                })
                
                # Создаем словарь для быстрого доступа
                charges_dict = {}
                for _, row in df_renamed.iterrows():
                    msisdn = str(row['MSISDN_charges']).strip()
                    try:
                        amount = float(row['AP_POSLE'])
                        charges_dict[msisdn] = amount
                    except:
                        charges_dict[msisdn] = 0.0
                        
                # Сохраняем в кеш
                self.cache['charges_df'] = df_renamed
                self.cache['charges_dict'] = charges_dict
                self.cache['charges_file'] = file_path
                self.cache['charges_columns'] = charges_mapping
                self.cache['charges_loaded'] = True
                self.cache['charges_count'] = len(charges_dict)
                self.cache['current_mrf'] = current_mrf
                
                self.files_loaded['charges'] = file_path
                self.update_files_info()
                self.update_cache_status()
                self.cache_mrf_label.setText(f"Текущий МРФ: {current_mrf}")
                
                self.add_log(f"Загружены начисления для МРФ {current_mrf}")
                self.add_log(f"  Файл: {os.path.basename(file_path)}")
                self.add_log(f"  Количество записей: {len(charges_dict):,}")
                self.add_log(f"  Колонки: MSISDN={charges_mapping['msisdn']}, Сумма={charges_mapping['charges']}")
                
                self.check_ready()
                
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить файл начислений:\n{str(e)}")
                self.add_log(f"Ошибка загрузки начислений: {str(e)}")
                self.clear_cache()
                
    def update_files_info(self):
        info = ""
        if self.files_loaded['report']:
            info += f"✅ Отчет: {os.path.basename(self.files_loaded['report'])}\n"
        else:
            info += "❌ Отчет: не загружен\n"
            
        if self.files_loaded['base']:
            for i, base in enumerate(self.files_loaded['base']):
                info += f"✅ База индексации {i+1}: {os.path.basename(base)}\n"
        else:
            info += "❌ База индексации: не загружена\n"
            
        if self.cache['charges_loaded']:
            info += f"✅ Начисления (кеш): {os.path.basename(self.cache['charges_file'])} "
            info += f"({self.cache['charges_count']:,} записей, МРФ: {self.cache['current_mrf']})\n"
        else:
            info += "❌ Начисления: не загружены\n"
            
        self.files_info.setText(info)
        
    def on_mode_changed(self):
        if self.mode_a_radio.isChecked():
            self.data['mode'] = 'A'
            self.load_charges_btn.setEnabled(True)
            self.add_log("Выбран режим А: начисления в отдельном файле")
        else:
            self.data['mode'] = 'B'
            self.load_charges_btn.setEnabled(False)
            self.add_log("Выбран режим Б: начисления в базе индексации")
            # При переключении на режим Б очищаем кеш
            if self.cache['charges_loaded']:
                self.clear_cache()
                self.add_log("Кеш очищен при смене режима на Б")
        self.check_ready()
        
    def setup_columns(self):
        """Настройка соответствия колонок в базе индексации"""
        if self.data['base_df'] is None:
            QMessageBox.warning(self, "Ошибка", "Сначала загрузите базу индексации!")
            return
            
        columns = list(self.data['base_df'].columns)
        
        dialog = ColumnSelectionDialog(columns, "Выбор колонок в базе индексации", self)
        if dialog.exec_() == QDialog.Accepted:
            selections = dialog.get_selections()
            if selections:
                self.data['column_mapping'] = selections
                info_text = "Настроены колонки:\n"
                for key, value in selections.items():
                    info_text += f"  {key}: {value}\n"
                self.column_info.setText(info_text)
                self.column_info.setStyleSheet("color: green;")
                self.add_log(f"Настроены колонки: {selections}")
                self.check_ready()
                
                # Если есть процент в базе, автоматически заполняем
                if 'percent' in selections and selections['percent']:
                    try:
                        percent_val = self.data['base_df'][selections['percent']].iloc[0]
                        if pd.notna(percent_val):
                            self.percent_edit.setText(str(percent_val))
                    except:
                        pass
            else:
                QMessageBox.warning(self, "Предупреждение", "Не выбрано ни одной колонки!")
                
    def select_date(self):
        """Выбор даты индексации"""
        dialog = DateSelectionDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            date = dialog.get_date()
            if date:
                self.data['indexation_date'] = date
                self.date_label.setText(date.strftime("%d.%m.%Y"))
                self.add_log(f"Выбрана дата индексации: {date.strftime('%d.%m.%Y')}")
                self.check_ready()
                
    def check_ready(self):
        """Проверяет готовность к запуску"""
        ready = True
        issues = []
        
        if not self.data['selected_mrf']:
            ready = False
            issues.append("Не выбран ни один МРФ")
            
        if self.files_loaded['report'] is None:
            ready = False
            issues.append("Не загружен файл отчета")
            
        if not self.files_loaded['base']:
            ready = False
            issues.append("Не загружена база индексации")
            
        if self.data['mode'] == 'A':
            if not self.cache['charges_loaded']:
                ready = False
                issues.append("Не загружены начисления для текущего МРФ")
            elif self.cache['current_mrf'] not in self.data['selected_mrf']:
                ready = False
                issues.append(f"Начисления загружены для другого МРФ: {self.cache['current_mrf']}")
        
        if self.data['selected_month'] is None:
            ready = False
            issues.append("Не выбран отчетный месяц")
            
        if not self.data['column_mapping']:
            ready = False
            issues.append("Не настроены колонки в базе индексации")
            
        if self.data['indexation_date'] is None:
            ready = False
            issues.append("Не выбрана дата индексации")
            
        if self.service_combo.currentIndex() == 0:
            ready = False
            issues.append("Не выбрана услуга")
            
        self.run_btn.setEnabled(ready)
        
        if ready:
            self.run_btn.setToolTip("Готов к запуску")
            self.data['selected_service'] = self.service_combo.currentText()
            self.data['indexation_percent'] = self.percent_edit.text()
            
            # Получаем выбранные сегменты
            selected_short = [item.text() for item in self.short_seg_list.selectedItems()]
            selected_sub = [item.text() for item in self.sub_seg_list.selectedItems()]
            self.data['selected_segments'] = selected_short
            self.data['selected_subsegments'] = selected_sub
        else:
            self.run_btn.setToolTip("Не готов: " + "; ".join(issues))
            
    def run_calculation(self):
        try:
            self.add_log("=" * 60)
            self.add_log("НАЧАЛО РАСЧЕТА")
            self.progress_bar.setValue(0)
            
            # Получаем данные
            base_df = self.data['base_df'].copy()
            mapping = self.data['column_mapping']
            
            # Проверяем наличие необходимых колонок
            required = ['msisdn', 'charges_before', 'mrf', 'region']
            for req in required:
                if req not in mapping:
                    QMessageBox.warning(self, "Ошибка", f"Не выбрана колонка: {req}")
                    return
                    
            # Переименовываем колонки для удобства
            rename_map = {
                mapping['msisdn']: 'MSISDN',
                mapping['charges_before']: 'AP_DO',
                mapping['mrf']: 'МРФ_исх',
                mapping['region']: 'РФ_исх'
            }
            
            if 'short_seg' in mapping:
                rename_map[mapping['short_seg']] = 'Краткий_сегмент'
            if 'sub_seg' in mapping:
                rename_map[mapping['sub_seg']] = 'Подсегмент'
            if 'percent' in mapping:
                rename_map[mapping['percent']] = 'Процент_индексации'
                
            base_df = base_df.rename(columns=rename_map)
            
            # Получаем выбранные МРФ
            selected_mrf = self.data['selected_mrf']
            month = self.data['selected_month']
            service = self.data['selected_service']
            index_date = self.data['indexation_date']
            
            self.add_log(f"Услуга: {service}")
            self.add_log(f"Дата индексации: {index_date.strftime('%d.%m.%Y')}")
            self.add_log(f"Отчетный месяц: {month}")
            self.add_log(f"Выбрано МРФ: {', '.join(selected_mrf)}")
            
            # Обрабатываем каждый МРФ
            total_processed = 0
            total_mrf = len(selected_mrf)
            
            for i, mrf in enumerate(selected_mrf):
                self.progress_bar.setValue(int((i / total_mrf) * 40))
                self.add_log(f"\n{'='*40}")
                self.add_log(f"Обработка МРФ: {mrf} ({i+1}/{total_mrf})")
                
                # Обновляем текущий МРФ
                self.data['processing_mrf'] = mrf
                
                # В режиме А запрашиваем начисления для этого МРФ
                if self.data['mode'] == 'A':
                    # Проверяем, есть ли кеш для этого МРФ
                    if not self.cache['charges_loaded'] or self.cache['current_mrf'] != mrf:
                        self.add_log(f"  Запрашиваем начисления для МРФ {mrf}")
                        reply = QMessageBox.question(
                            self,
                            "Загрузка начислений",
                            f"Для МРФ {mrf} не загружены начисления.\n"
                            "Загрузить файл начислений?",
                            QMessageBox.Yes | QMessageBox.No
                        )
                        if reply == QMessageBox.No:
                            self.add_log(f"  Пропускаем МРФ {mrf} (нет начислений)")
                            continue
                        else:
                            # Очищаем старый кеш и загружаем новый
                            self.clear_cache()
                            self.load_charges_file()
                            if not self.cache['charges_loaded']:
                                self.add_log(f"  Пропускаем МРФ {mrf} (не удалось загрузить начисления)")
                                continue
                    
                    # Используем кешированные начисления
                    charges_dict = self.cache['charges_dict']
                    self.add_log(f"  Используем кешированные начисления из файла: {os.path.basename(self.cache['charges_file'])}")
                    self.add_log(f"    Записей в кеше: {len(charges_dict):,}")
                    
                    # Применяем словарь к базе
                    base_df_copy = base_df.copy()
                    base_df_copy['AP_POSLE'] = base_df_copy['MSISDN'].astype(str).map(charges_dict).fillna(0)
                    
                else:  # Режим Б
                    # Пользователь выбирает колонку с начислениями ПОСЛЕ
                    self.add_log(f"  Запрашиваем колонку с начислениями ПОСЛЕ для МРФ {mrf}")
                    columns = list(base_df.columns)
                    dialog = ColumnSelectionDialog(columns, 
                        f"Выбор колонки с начислениями ПОСЛЕ для МРФ {mrf}", 
                        self, 
                        required_only=True
                    )
                    if dialog.exec_() != QDialog.Accepted:
                        self.add_log(f"  Пропускаем МРФ {mrf} (не выбрана колонка)")
                        continue
                        
                    charges_mapping = dialog.get_selections()
                    if 'charges' not in charges_mapping:
                        self.add_log(f"  Пропускаем МРФ {mrf} (не выбрана колонка с начислениями)")
                        continue
                        
                    base_df_copy = base_df.copy()
                    base_df_copy = base_df_copy.rename(columns={
                        charges_mapping['charges']: 'AP_POSLE'
                    })
                
                # Получаем регионы для МРФ
                regions = MRF_TO_RF.get(mrf, [])
                
                # Фильтруем данные по МРФ
                mrf_data = base_df_copy[base_df_copy['МРФ_исх'].str.contains(mrf, case=False, na=False)]
                
                if mrf_data.empty:
                    self.add_log(f"  Нет данных для МРФ {mrf}")
                    continue
                    
                self.add_log(f"  Найдено записей в базе: {len(mrf_data):,}")
                
                # Проверяем наличие сегментов
                has_short_seg = 'Краткий_сегмент' in mrf_data.columns
                has_sub_seg = 'Подсегмент' in mrf_data.columns
                
                # Рассчитываем эффект по каждому региону
                mrf_total = 0
                for region in regions:
                    region_data = mrf_data[mrf_data['РФ_исх'].str.contains(region, case=False, na=False)]
                    if region_data.empty:
                        continue
                        
                    # Группируем по сегментам
                    if has_short_seg and has_sub_seg:
                        grouped = region_data.groupby(['Краткий_сегмент', 'Подсегмент']).agg({
                            'AP_DO': 'sum',
                            'AP_POSLE': 'sum'
                        })
                        
                        for (short_seg, sub_seg), row in grouped.iterrows():
                            effect = max(0, row['AP_POSLE'] - row['AP_DO'])
                            if effect > 0:
                                self.write_to_report(
                                    mrf=mrf,
                                    region=region,
                                    short_seg=short_seg,
                                    sub_seg=sub_seg,
                                    effect=effect / 1000,
                                    service=service,
                                    index_date=index_date,
                                    month=month
                                )
                                total_processed += 1
                                mrf_total += 1
                    else:
                        # Если сегментов нет - группируем по региону
                        total_after = region_data['AP_POSLE'].sum()
                        total_before = region_data['AP_DO'].sum()
                        effect = max(0, total_after - total_before)
                        
                        if effect > 0:
                            self.write_to_report(
                                mrf=mrf,
                                region=region,
                                short_seg='Все',
                                sub_seg='Все',
                                effect=effect / 1000,
                                service=service,
                                index_date=index_date,
                                month=month
                            )
                            total_processed += 1
                            mrf_total += 1
                
                self.add_log(f"  Записано записей для МРФ {mrf}: {mrf_total}")
                
                # Очищаем кеш после обработки МРФ (если это не последний МРФ)
                if i < len(selected_mrf) - 1:
                    self.add_log(f"  Очищаем кеш начислений после обработки МРФ {mrf}")
                    self.clear_cache()
                    self.cache_mrf_label.setText("Текущий МРФ: не выбран")
                    self.add_log("  Готов к следующему МРФ")
                
                self.progress_bar.setValue(int(((i + 1) / total_mrf) * 80))
                    
            self.progress_bar.setValue(100)
            self.add_log(f"\n{'='*60}")
            self.add_log(f"✅ РАСЧЕТ ЗАВЕРШЕН")
            self.add_log(f"Обработано записей: {total_processed}")
            self.add_log("=" * 60)
            
            QMessageBox.information(
                self, 
                "Готово", 
                f"Расчет успешно завершен!\n"
                f"Обработано записей: {total_processed}\n"
                f"Результаты записаны в файл отчета."
            )
            
            self.show_results()
            
        except Exception as e:
            # При ошибке очищаем кеш
            self.add_log(f"ОШИБКА: {str(e)}")
            self.clear_cache()
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка при расчете:\n{str(e)}")
            import traceback
            self.add_log(traceback.format_exc())
            
    def write_to_report(self, mrf, region, short_seg, sub_seg, effect, service, index_date, month):
        """Записывает результат в файл отчета"""
        try:
            report_path = self.files_loaded['report']
            wb = load_workbook(report_path)
            
            sheet_name = 'Факт-эффект индексации 2026'
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
            ws = wb[sheet_name]
            
            # Определяем заголовки
            headers = []
            for row in ws.iter_rows(min_row=1, max_row=1):
                headers = [cell.value for cell in row]
                break
                
            if not headers:
                headers = ['Услуга', 'Дата индексации', 'Процент индексации', 'МРФ', 'РФ', 
                          'Краткий сегмент', 'Подсегмент', 
                          f'ФАКТ Эффект от индексации тыс.руб.без НДС {month} 2026']
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_idx, value=header)
                    
            # Находим колонку для записи эффекта
            month_col_pattern = f'ФАКТ Эффект от индексации тыс.руб.без НДС {month} 2026'
            effect_col = None
            
            for col_idx, header in enumerate(headers, 1):
                if header and month_col_pattern in str(header):
                    effect_col = col_idx
                    break
                    
            if effect_col is None:
                effect_col = len(headers) + 1
                ws.cell(row=1, column=effect_col, value=month_col_pattern)
                
            # Маппинг МРФ для отчета
            mrf_map = {
                'Северо-Запад': 'СЗ',
                'Волга': 'Волга',
                'Дальний Восток': 'ДВ',
                'Урал': 'Урал',
                'Центр': 'Центр',
                'Юг': 'Юг',
                'Сибирь': 'Сибирь'
            }
            
            # Стили
            green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
            blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
            
            # Ищем существующую строку
            found = False
            for row_idx in range(2, ws.max_row + 1):
                mrf_cell = ws.cell(row=row_idx, column=4).value
                region_cell = ws.cell(row=row_idx, column=5).value
                short_cell = ws.cell(row=row_idx, column=6).value
                sub_cell = ws.cell(row=row_idx, column=7).value
                
                if (mrf_cell == mrf_map.get(mrf, mrf) and 
                    region_cell == region and
                    short_cell == short_seg and
                    sub_cell == sub_seg):
                    
                    current_value = ws.cell(row=row_idx, column=effect_col).value or 0
                    ws.cell(row=row_idx, column=effect_col, value=current_value + effect)
                    ws.cell(row=row_idx, column=effect_col).fill = green_fill
                    found = True
                    break
                    
            if not found:
                # Добавляем новую строку
                new_row = ws.max_row + 1
                ws.cell(row=new_row, column=1, value=service)
                ws.cell(row=new_row, column=2, value=index_date.strftime('%d.%m.%Y'))
                ws.cell(row=new_row, column=3, value=self.percent_edit.text())
                ws.cell(row=new_row, column=4, value=mrf_map.get(mrf, mrf))
                ws.cell(row=new_row, column=5, value=region)
                ws.cell(row=new_row, column=6, value=short_seg)
                ws.cell(row=new_row, column=7, value=sub_seg)
                ws.cell(row=new_row, column=effect_col, value=effect)
                ws.cell(row=new_row, column=effect_col).fill = blue_fill
                
            wb.save(report_path)
            
        except Exception as e:
            self.add_log(f"  Ошибка при записи в отчет для {region}: {str(e)}")
            raise
            
    def show_results(self):
        """Показывает результаты"""
        try:
            if self.files_loaded['report']:
                df = pd.read_excel(self.files_loaded['report'], sheet_name='Факт-эффект индексации 2026')
                self.results_text.setText(df.tail(20).to_string())
                self.add_log(f"Показаны последние 20 записей из отчета")
        except Exception as e:
            self.results_text.setText(f"Не удалось отобразить результаты: {str(e)}")
            
    def closeEvent(self, event):
        """Обработка закрытия программы - очищаем кеш"""
        if self.cache['charges_loaded']:
            self.add_log(f"Очистка кеша начислений при закрытии программы")
            self.clear_cache()
        event.accept()

def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    
    # Установка иконки приложения
    app.setWindowIcon(QIcon())
    
    window = IndexationApp()
    window.show()
    
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()